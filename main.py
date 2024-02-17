from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import random
from pprint import pprint


def initialize():
    # Initialize global variables
    global students, timeSlots, organizations, orgNameToCode, timeCodeToSlot
    global dataExcelFile, processingWorkbook, studentPrefSheet, processingSheet
    global maxPreferencesNo

    # Dictionaries for storing data
    students = {}
    orgNameToCode = {}
    timeCodeToSlot = {}
    organizations = {}

    maxPreferencesNo = 0

    # Load the workbook containing student and organization data
    dataExcelFile = load_workbook("DataFiles/StudentResidencyData.xlsx")
    orgSheet = dataExcelFile["OrganizationDetails"]
    studentPrefSheet = dataExcelFile["StudentPreferenceDetails"]
    timeSlotSheet = dataExcelFile["TimeSlotDetails"]

    # Create a new workbook for processing the allocations
    processingWorkbook = Workbook()
    processingWorkbook.create_sheet("StudentPreferences", 0)
    processingSheet = processingWorkbook["StudentPreferences"]

    # Initialize organization name to code mapping
    for i in range(2, orgSheet.max_row + 1):
        orgName = orgSheet.cell(i, 1).value
        orgNameToCode[orgName] = {'code': orgSheet.cell(i, 2).value}

    # Initialize time code to slot mapping
    timeSlots = timeSlotSheet.max_row - 1
    for i in range(2, timeSlotSheet.max_row + 1):
        timeCode = timeSlotSheet.cell(i, 1).value
        timeCodeToSlot[timeCode] = {'slot': timeSlotSheet.cell(i, 2).value}

    # Initialize organizations with slots and students per slot
    for i in range(2, orgSheet.max_row + 1):
        orgCode = orgSheet.cell(i, 2).value
        organizations[orgCode] = {
            'name': orgSheet.cell(i, 1).value,
            'slotsAllocatedToOrg': orgSheet.cell(i, 3).value,
            'studentsPerSlot': orgSheet.cell(i, 4).value,  # Students per slot
            'allocatedStudents': 0,
            'studentsIDSlotMapping': {}
        }
        for j in range(1, organizations[orgCode]['slotsAllocatedToOrg']+1):
            organizations[orgCode]['studentsIDSlotMapping'][j] = []
 
    # Initialize students with preferences
    for i in range(2, studentPrefSheet.max_row + 1):
        uscId = studentPrefSheet.cell(i, 2).value
        students[uscId] = {
            'name': studentPrefSheet.cell(i, 1).value,
            'allocatedOrganizations': 0,
            'organizationsCodeSlotMapping': {},
            "preferences": {}
        }
        for slot in range(1, timeSlots+1):
            students[uscId]['organizationsCodeSlotMapping'][slot] = None

        # New logic for reading preferences based on the new Excel format
        for j in range(7, studentPrefSheet.max_column + 1):
            prefValue = studentPrefSheet.cell(i, j).value
            # Strip whitespace and check for non-empty string
            if prefValue is not None and str(prefValue).strip():
                orgCode = studentPrefSheet.cell(1, j).value.strip()
                if orgCode in orgNameToCode:
                    maxPreferencesNo = max(maxPreferencesNo, int(prefValue))
                    # Now we know prefValue is indeed a valid preference number and not an empty or whitespace-only cell
                    students[uscId]["preferences"][int(prefValue)] = orgNameToCode[orgCode]['code']
            
    # print("Initialization complete.")
    # print("Students:")
    # pprint(students)



def preprocessing_student_preferences_sheet():
    global students
    global orgNameToCode
    global processingSheet
    global studentPrefSheet

    # Copy the student preferences headings to the processing sheet
    processingSheet.cell(1, 1).value = "Student Name"
    processingSheet.cell(1, 1).font = Font(bold=True)
    processingSheet.cell(1, 2).value = "USC ID"
    processingSheet.cell(1, 2).font = Font(bold=True)


    for i in range(1, maxPreferencesNo + 1):
        processingSheet.cell(1, i+2).value = "Pref - "+str(i)
        processingSheet.cell(1, i+2).font = Font(bold=True)

    # Populate the processing sheet with student preferences from student dictionary preferencese list
    for i, (uscId, student) in enumerate(students.items(), start=2):
        pprint(student)
        processingSheet.cell(i, 1).value = student['name']
        processingSheet.cell(i, 2).value = uscId
        for pref, orgCode in student['preferences'].items():
            processingSheet.cell(i, pref+2).value = orgCode


def dynamic_allocation_of_students():
    # Dynamic allocation based on student preferences and organization capacity
    last_allocated_preference = {studentUSCId: 0 for studentUSCId in students}

    allocation_possible = True
    while allocation_possible:
        allocation_possible = False
        student_ids = list(students.keys())
        random.shuffle(student_ids)

        for studentUSCId in student_ids:
            student_preferences = students[studentUSCId]["preferences"]
            if students[studentUSCId]['allocatedOrganizations'] >= timeSlots:
                continue  # Student already allocated maximum organizations

            # Get the list of preference numbers sorted in ascending order
            sorted_preferences = sorted(student_preferences.keys())
            
            # Iterate through the student's preferences
            for pref_index in sorted_preferences:
                # Skip preferences that have already been allocated
                if pref_index <= last_allocated_preference[studentUSCId]:
                    continue

                currentOrganization = student_preferences[pref_index]
                if organizations.get(currentOrganization) is None:
                    continue  # Skip invalid preferences

                if organizations[currentOrganization]['allocatedStudents'] >= organizations[currentOrganization]['slotsAllocatedToOrg'] * organizations[currentOrganization]['studentsPerSlot']:
                    continue  # Organization at capacity

                for slot in range(1, timeSlots + 1):
                    if students[studentUSCId]['organizationsCodeSlotMapping'][slot] is not None:
                        continue  # Student already allocated to a slot

                    if len(organizations[currentOrganization]['studentsIDSlotMapping'][slot]) < organizations[currentOrganization]['studentsPerSlot']:
                        # Allocate student to this organization and slot
                        organizations[currentOrganization]['allocatedStudents'] += 1
                        students[studentUSCId]['allocatedOrganizations'] += 1
                        organizations[currentOrganization]['studentsIDSlotMapping'][slot].append(studentUSCId)
                        students[studentUSCId]['organizationsCodeSlotMapping'][slot] = currentOrganization
                        
                        last_allocated_preference[studentUSCId] = pref_index
                        allocation_possible = True  # Allocation was successful
                        break  # Break out of the slot loop, proceed to next student



def populate_processing_workbook():
    # Assuming processingWorkbook is a global variable or passed as an argument
    global processingWorkbook, timeCodeToSlot, students, organizations

    # Create sheets for the allocation results
    processingWorkbook.create_sheet("StudentsMapping", 0)
    studentsMappingSheet = processingWorkbook["StudentsMapping"]
    
    # Set headers for the StudentsMapping sheet
    studentsMappingSheet.cell(1, 1).value = "USC ID"
    studentsMappingSheet.cell(1, 2).value = "Student Name"
    studentsMappingSheet.cell(1, 1).font = Font(bold=True)
    studentsMappingSheet.cell(1, 2).font = Font(bold=True)

    # Add headers for organization slots
    for i, time_slot in enumerate(sorted(timeCodeToSlot.keys()), start=3):
        studentsMappingSheet.cell(1, i).value = timeCodeToSlot[time_slot]['slot']
        studentsMappingSheet.cell(1, i).font = Font(bold=True)

    # Populate the student allocations
    for row_index, (uscId, student) in enumerate(students.items(), start=2):
        studentsMappingSheet.cell(row_index, 1).value = uscId
        studentsMappingSheet.cell(row_index, 2).value = student['name']
        
        # Fill in organization allocations based on the student's slot mapping
        for slot, org_code in student['organizationsCodeSlotMapping'].items():
            col_index = list(sorted(timeCodeToSlot.keys())).index(slot) + 3
            if org_code:
                org_name = organizations[org_code]['name']
                studentsMappingSheet.cell(row_index, col_index).value = org_name

    # Create sheet for Organization Mapping
    processingWorkbook.create_sheet("OrganizationMapping", 1)
    organizationMappingSheet = processingWorkbook["OrganizationMapping"]

    # Set headers for the OrganizationMapping sheet
    organizationMappingSheet.cell(1, 1).value = "Organization Code"
    organizationMappingSheet.cell(1, 2).value = "Organization Name"
    organizationMappingSheet.cell(1, 1).font = Font(bold=True)
    organizationMappingSheet.cell(1, 2).font = Font(bold=True)

    # Add headers for organization slots
    for i, time_slot in enumerate(sorted(timeCodeToSlot.keys()), start=3):
        organizationMappingSheet.cell(1, i).value = timeCodeToSlot[time_slot]['slot']
        organizationMappingSheet.cell(1, i).font = Font(bold=True)

    # Populate the organization allocations
    for row_index, (org_code, org) in enumerate(organizations.items(), start=2):
        organizationMappingSheet.cell(row_index, 1).value = org_code
        organizationMappingSheet.cell(row_index, 2).value = org['name']
        
        # Fill in student names based on each organization's slot mapping
        for slot, student_ids in org['studentsIDSlotMapping'].items():
            col_index = list(sorted(timeCodeToSlot.keys())).index(slot) + 3
            student_names = [students[student_id]['name'] for student_id in student_ids]
            organizationMappingSheet.cell(row_index, col_index).value = ", ".join(student_names)

    # Save the workbook to a new file
    processingWorkbook.save("DataFiles/Final_Dynamic_Mapping.xlsx")

def main():
    initialize()
    preprocessing_student_preferences_sheet()
    dynamic_allocation_of_students()
    populate_processing_workbook()

if __name__ == "__main__":
    main()
